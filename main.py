import tkinter as tk
from tkinter import filedialog, messagebox
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Alignment
import os

def select_file():
    file_path = filedialog.askopenfilename(filetypes=[("HTML files", "*.html")])
    job_number = job_number_entry.get()
    if file_path and job_number:
        try:
            process_file(file_path, job_number)
            messagebox.showinfo("Success", "Excel file updated successfully.")
            root.destroy()  # Close the GUI after success
        except Exception as e:
            messagebox.showerror("Error", f"An error occurred: {e}")

def parse_html(file_path):
    with open(file_path, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')

    # Define the keys to extract
    keys_to_extract = [
        'ACM hardware class', 'ACM version', 'ACM diagnosis version', 'ACM VIN', 'ACM serial number',
        'ACM hardware part number', 'ACM certification', 'ACM hardware version'
    ]

    # Extract the values
    extracted_values = {key: None for key in keys_to_extract}

    rows = soup.find_all('tr')
    for row in rows:
        cells = row.find_all('td')
        if len(cells) == 2:
            key = cells[0].get_text(strip=True)
            value = cells[1].get_text(strip=True)
            if key == 'ACM diagnosis version':
                value = value.lstrip('0')  # Remove leading zeros
            if key in extracted_values:
                extracted_values[key] = value

    # Debug print to check extracted data
    for key, value in extracted_values.items():
        print(f"{key}: {value}")

    return extracted_values

def update_excel(extracted_values, job_number, excel_path):
    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f"Excel file not found: {excel_path}")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active  # Or specify the sheet name: wb['SheetName']

    # Header mapping: Map the keys to the actual column headers in the Excel file
    header_mapping = {
        'ACM hardware class': 'Hardware Class',
        'ACM version': 'Version',
        'ACM diagnosis version': 'Diagnosis Version',
        'ACM VIN': 'Vin',
        'ACM serial number': 'Serial Number',
        'ACM hardware part number': 'Part Number',
        'ACM certification': 'Certification',
        'ACM hardware version': 'Hardware Version',
        'Job number': 'Fixably No.'
    }

    # Identify the header row
    header_row_index = None
    for row in ws.iter_rows(min_row=1, max_row=10):  # Adjust range if headers are farther down
        headers = {cell.value: cell.column for cell in row if cell.value}
        print(f"Headers found in row {row[0].row}: {headers}")  # Debug print
        if set(header_mapping.values()).issubset(headers.keys()):
            header_row_index = row[0].row
            break

    if not header_row_index:
        raise ValueError("Header row not found in the Excel sheet")

    # Find the correct column indices based on headers
    headers = {cell.value: cell.column for cell in ws[header_row_index]}
    print(f"Headers and their columns: {headers}")  # Debug print

    # Check if all required columns are present
    for key in extracted_values.keys():
        if header_mapping[key] not in headers:
            raise ValueError(f"Column for '{key}' not found in the Excel sheet")

    # Add job number to the extracted values
    extracted_values['Job number'] = job_number

    # Find the lowest ID number with an otherwise empty row
    target_row = None
    for row in ws.iter_rows(min_row=header_row_index + 1):
        id_cell = row[0]  # Assuming ID Number is in the first column
        if id_cell.value is not None:
            # Check if the rest of the cells in this row are empty
            if all(cell.value is None for cell in row if cell.column != 1):  # Exclude ID Number column
                target_row = id_cell.row
                break

    if target_row is None:
        raise ValueError("No suitable row found for updating")

    # Insert values into the identified row
    for key, value in extracted_values.items():
        cell = ws.cell(row=target_row, column=headers[header_mapping[key]], value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center')  # Center-align the cell

    wb.save(excel_path)
    wb.close()

def process_file(file_path, job_number):
    extracted_values = parse_html(file_path)
    excel_path = r'C:\Users\User\OneDrive\Documents\ACM2_2.1\Unit Assessment ACM--PC.xlsx'
    update_excel(extracted_values, job_number, excel_path)

# GUI Setup
root = tk.Tk()
root.title("HTML to Excel")

frame = tk.Frame(root, padx=10, pady=10)
frame.pack(padx=10, pady=10)

job_number_label = tk.Label(frame, text="Job number:")
job_number_label.pack()

job_number_entry = tk.Entry(frame)
job_number_entry.pack()

select_button = tk.Button(frame, text="Select fault report", command=select_file)
select_button.pack()

root.mainloop()
